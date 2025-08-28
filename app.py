# app.py
import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
import streamlit as st
from lxml import etree
from openpyxl.utils import get_column_letter  # p/ formatar colunas no Excel

# parsers/__init__.py deve expor: NFe, NFCe, NFSe ABRASF, Evento NFe, NFSe RN (Prestado/Tomado), CT-e (se tiver)
from parsers import ALL_PARSERS, get_parser_by_name
from utils.io import iter_xml_paths_from_dir

# ---------------------------
# Config da página
# ---------------------------
st.set_page_config(page_title="Leitor de XML de Notas → Excel", layout="wide")

# ---- Estado (cache leve) ----
if "df" not in st.session_state:
    st.session_state.df = None
    st.session_state.df_view = None
    st.session_state.erros = None
    st.session_state.paths = []

# --- CSS/Estilo ---
st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
.block-container {padding-top: 1.5rem; padding-bottom: 3rem; max-width: 1280px;}
h1, h2, h3 { font-weight: 700; letter-spacing: -0.02em; }
.az-card {
  background: var(--secondary-background-color, #f6f8fc);
  border: 1px solid rgba(0,0,0,0.05);
  border-radius: 16px;
  padding: 1rem 1.1rem;
  box-shadow: 0 1px 2px rgba(0,0,0,0.04);
  margin-bottom: .75rem;
}
.stButton>button {
  border-radius: 10px;
  padding: .6rem 1rem;
  box-shadow: 0 1px 2px rgba(0,0,0,.06);
}
.badge {
  display: inline-block;
  padding: .25rem .6rem;
  border-radius: 9999px;
  background: #e5edff;
  color: #1e40af;
  font-size: .8rem;
  font-weight: 600;
  margin-right: .4rem;
  border: 1px solid rgba(37,99,235, .15);
}
[data-testid="stDataFrame"] table { font-size: 0.92rem; }
</style>
""", unsafe_allow_html=True)

# ---------------------------
# Cabeçalho (hero)
# ---------------------------
st.title("📄 Leitor de XML de Notas Fiscais → Excel")

col_logo, col_head = st.columns([1, 6], vertical_alignment="center")
with col_logo:
    from pathlib import Path as _Path
    logo_path = _Path("assets/logo.png")
    if logo_path.exists():
        st.image(str(logo_path), width=72)
with col_head:
    st.markdown("""
    
    """, unsafe_allow_html=True)

# ---------------------------
# UI - Configuração (card)
# ---------------------------
st.markdown("### ⚙️ Configuração")
st.markdown('<div class="az-card">', unsafe_allow_html=True)

tipos = ["Auto (detectar)"] + [p.name for p in ALL_PARSERS]
tipo = st.selectbox("Tipo de XML", tipos)

colA, colB = st.columns(2)
with colA:
    uploaded_files = st.file_uploader(
        "Upload de XMLs (múltiplos)", type=["xml"], accept_multiple_files=True
    )
with colB:
    dir_path = st.text_input(
        "Ou informe um diretório local com XMLs (recomendado para lotes enormes)",
        help="Ex.: C:\\\\Users\\\\seu.usuario\\\\notas_xml  ou  /dados/notas"
    )

colC, colD = st.columns([1,1])
with colC:
    max_workers = st.slider(
        "Paralelismo (threads)",
        min_value=4, max_value=64, value=64,step=4,
        help="Ajuste conforme CPU e armazenamento."
    )
with colD:
    inclui_eventos = st.checkbox("Incluir eventos (procEventoNFe) na tabela principal", value=True)

processar = st.button("Processar")
st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# Helpers
# ---------------------------
NFE_NS = "http://www.portalfiscal.inf.br/nfe"
CTE_NS = "http://www.portalfiscal.inf.br/cte"
ABRASF_NS = "http://www.abrasf.org.br/ABRASF/arquivos/nfse.xsd"

def cnpj_from_chave(chave_norm: Optional[str]) -> Optional[str]:
    """Extrai o CNPJ do emitente da chave NF-e (44 dígitos)."""
    if not chave_norm:
        return None
    s = normalize_key(chave_norm)
    if s and len(s) == 44:
        # cUF(2) + AAMM(4) = 6; próximos 14 dígitos = CNPJ emitente
        return s[6:20]
    return None


def normalize_key(val) -> Optional[str]:
    """Mantém apenas dígitos e retorna os últimos 44. Se não der, retorna None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = re.sub(r"\D", "", str(val))
    if len(s) < 44:
        return None
    return s[-44:]

def detect_parser(root: etree._Element):
    for p in ALL_PARSERS:
        try:
            if p.matches(root):
                return p
        except Exception:
            pass
    return None

def parse_with_selected_or_auto(root: etree._Element, nome_arquivo_hint: str, tipo_ui: str) -> Dict[str, Any]:
    """
    Tenta usar o parser selecionado. Se não casar, faz fallback para detecção automática.
    Assim, mesmo que o usuário escolha 'NF-e' e o arquivo seja 'Evento', o arquivo é lido.
    """
    parser_local = None
    if tipo_ui != "Auto (detectar)":
        sel = get_parser_by_name(tipo_ui)
        try:
            if sel.matches(root):
                parser_local = sel
            else:
                parser_local = detect_parser(root)
        except Exception:
            parser_local = detect_parser(root)
    else:
        parser_local = detect_parser(root)

    if not parser_local:
        raise ValueError("Nenhum parser reconheceu este XML.")

    data = parser_local.parse_header(root)
    data["_arquivo"] = nome_arquivo_hint
    data["_parser"] = parser_local.name
    return data

def parse_path(p: Path, tipo_ui: str) -> Dict[str, Any]:
    with open(p, "rb") as f:
        tree = etree.parse(f)
    root = tree.getroot()
    return parse_with_selected_or_auto(root, str(p), tipo_ui)

def parse_buffer_bytes(raw: bytes, name: str, tipo_ui: str) -> Dict[str, Any]:
    tree = etree.parse(io.BytesIO(raw))
    root = tree.getroot()
    return parse_with_selected_or_auto(root, name, tipo_ui)

def infer_movimento(row: Dict[str, Any]) -> str:
    tp = row.get("tpNF")
    if tp is None:
        return "Desconhecido"
    tp = str(tp).strip()
    if tp == "1":
        return "Saída"
    if tp == "0":
        return "Entrada"
    return "Desconhecido"

def to_number_maybe_br(x):
    if pd.isna(x):
        return pd.NA
    s = str(x).strip()
    if s == "":
        return pd.NA
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    return pd.to_numeric(s, errors="coerce")

def to_percent_decimal(x):
    v = to_number_maybe_br(x)
    if pd.isna(v):
        return pd.NA
    try:
        v = float(v)
    except Exception:
        return pd.NA
    return v / 100.0 if v > 1.0 else v

def to_datetime_col(x):
    return pd.to_datetime(x, errors="coerce")

def strip_tz_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    tz_cols = out.select_dtypes(include=["datetimetz"]).columns
    for c in tz_cols:
        try:
            out[c] = out[c].dt.tz_convert("America/Fortaleza").dt.tz_localize(None)
        except Exception:
            out[c] = out[c].dt.tz_localize(None)
    return out

def text_or_none(node: Optional[etree._Element], tag: str, ns: str) -> Optional[str]:
    if node is None:
        return None
    el = node.find(f"{{{ns}}}{tag}")
    if el is not None and el.text:
        return el.text.strip()
    return None

def sniff_minimal_from_bytes(raw: bytes) -> Dict[str, Any]:
    info: Dict[str, Any] = {}
    try:
        root = etree.fromstring(raw)
    except Exception as e:
        return {"_sniff_ok": False, "_sniff_erro": f"XML inválido: {e}"}

    try:
        nfe = root.find(f".//{{{NFE_NS}}}NFe")
        if nfe is None:
            nfe = root
        infNFe = nfe.find(f".//{{{NFE_NS}}}infNFe")
        if infNFe is not None:
            ide = infNFe.find(f".//{{{NFE_NS}}}ide")
            info["chave"] = (infNFe.get("Id") or "").replace("NFe", "") or None
            info["modelo"] = text_or_none(ide, "mod", NFE_NS)
            info["tpAmb"] = text_or_none(ide, "tpAmb", NFE_NS)
            info["nNF"] = text_or_none(ide, "nNF", NFE_NS)
            info["emissao"] = text_or_none(ide, "dhEmi", NFE_NS) or text_or_none(ide, "dEmi", NFE_NS)
            info["_sniff_tipo"] = "NFe/NFCe"
            info["_sniff_ok"] = True
            return info
    except Exception:
        pass

    try:
        cte = root.find(f".//{{{CTE_NS}}}CTe")
        if cte is None:
            cte = root
        infCte = cte.find(f".//{{{CTE_NS}}}infCte")
        if infCte is not None:
            ide = infCte.find(f".//{{{CTE_NS}}}ide")
            info["chave"] = (infCte.get("Id") or "").replace("CTe", "") or None
            info["modelo"] = text_or_none(ide, "mod", CTE_NS)
            info["tpAmb"] = text_or_none(ide, "tpAmb", CTE_NS)
            info["nCT"] = text_or_none(ide, "nCT", CTE_NS)
            info["emissao"] = text_or_none(ide, "dhEmi", CTE_NS)
            info["_sniff_tipo"] = "CTe"
            info["_sniff_ok"] = True
            return info
    except Exception:
        pass

    try:
        inf = root.find(f".//{{{ABRASF_NS}}}InfNfse")
        if inf is not None:
            info["numero"] = text_or_none(inf, "Numero", ABRASF_NS)
            info["emissao"] = text_or_none(inf, "DataEmissao", ABRASF_NS)
            info["competencia"] = text_or_none(inf, "Competencia", ABRASF_NS)
            info["_sniff_tipo"] = "NFSe"
            info["_sniff_ok"] = True
            return info
    except Exception:
        pass

    return {"_sniff_ok": False}

# ---------------------------
# Processamento (somente se clicou)
# ---------------------------
if processar:
    # Coletar arquivos
    paths: List[Path] = []
    if dir_path.strip():
        try:
            base = Path(dir_path.strip())
            if not base.exists():
                st.error("O diretório informado não existe.")
                st.stop()
            paths = list(iter_xml_paths_from_dir(str(base)))
            paths = list(dict.fromkeys([Path(p).resolve() for p in paths]))  # dedup
        except Exception as e:
            st.error(f"Erro ao varrer o diretório: {e}")
            st.stop()

    mem_buffers = uploaded_files or []
    total_estimado = len(paths) + len(mem_buffers)
    if total_estimado == 0:
        st.warning("Forneça arquivos (upload) ou um diretório.")
        st.stop()

    st.info(f"Arquivos estimados: **{total_estimado}**")
    progress = st.progress(0)
    status_area = st.empty()

    results: List[Dict[str, Any]] = []
    erros: List[Dict[str, Any]] = []
    processed = 0
    future_ctx: Dict[Any, Dict[str, Any]] = {}

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for p in paths:
            fut = ex.submit(parse_path, p, tipo)
            future_ctx[fut] = {"src_type": "path", "name": str(p), "raw": None, "tipo_ui": tipo}

        for b in mem_buffers:
            name = getattr(b, "name", "uploaded.xml")
            raw = b.getvalue()
            fut = ex.submit(parse_buffer_bytes, raw, name, tipo)
            future_ctx[fut] = {"src_type": "upload", "name": name, "raw": raw, "tipo_ui": tipo}

        for fut in as_completed(list(future_ctx.keys())):
            meta = future_ctx[fut]
            try:
                row = fut.result()
                results.append(row)
            except Exception as e:
                try:
                    if meta["raw"] is not None:
                        sniff = sniff_minimal_from_bytes(meta["raw"])
                    else:
                        raw_path = Path(meta["name"]).read_bytes()
                        sniff = sniff_minimal_from_bytes(raw_path)
                except Exception as e2:
                    sniff = {"_sniff_ok": False, "_sniff_erro": f"Falha ao ler/inspecionar: {e2}"}

                errrow = {"_arquivo": meta["name"], "_parser_ui": meta["tipo_ui"], "_erro": str(e)}
                errrow.update(sniff)
                erros.append(errrow)
            finally:
                processed += 1
                pct = int(processed / total_estimado * 100)
                progress.progress(pct)
                status_area.info(f"Processados: {processed}/{total_estimado} ({pct}%)")

    df = pd.DataFrame(results)

    # --- Normalizações gerais / enriquecimento antes do cancelamento ---
    if not df.empty:
        # Moedas
        if "vNF" in df.columns: df["vNF"] = df["vNF"].map(to_number_maybe_br)
        if "valor_iss" in df.columns: df["valor_iss"] = df["valor_iss"].map(to_number_maybe_br)
        for col in ["vTPrest", "vRec", "vCarga"]:
            if col in df.columns: df[col] = df[col].map(to_number_maybe_br)

        # Datas
        for dcol in ["emissao", "competencia", "cancelado_em"]:
            if dcol in df.columns: df[dcol] = df[dcol].map(to_datetime_col)

        # Alíquota base 1
        if "aliquota" in df.columns: df["aliquota"] = df["aliquota"].map(to_percent_decimal)

        # Movimento (NF-e/NFC-e)
        if "tpNF" in df.columns:
            df["movimento"] = df.apply(lambda r: infer_movimento(r.to_dict()), axis=1)
        else:
            df["movimento"] = df.get("movimento", "Desconhecido")

    # --- CANCELAMENTO 110111 ⇒ ENVIAR PARA "ERROS" e EXCLUIR DA TABELA PRINCIPAL ---
    if len(df) > 0:
        # 0) __key (chave normalizada) p/ todos
        df["__key"] = None
        if "chave" in df.columns:
            df.loc[df["_parser"].isin(["NF-e", "NFC-e"]), "__key"] = df.loc[df["_parser"].isin(["NF-e", "NFC-e"]), "chave"].map(normalize_key)
        if "chNFe" in df.columns:
            mask_ev = df["_parser"].eq("Evento NF-e")
            df.loc[mask_ev & df["__key"].isna(), "__key"] = df.loc[mask_ev, "chNFe"].map(normalize_key)

        # 1) lookup do último evento de cancelamento por chave
        ev = df[df["_parser"] == "Evento NF-e"].copy()
        cancel_info = pd.DataFrame(columns=["__key", "cancelado_em", "cancel_nProt", "_arquivo_evento"])
        if not ev.empty:
            if "tpEvento" in ev.columns:
                ev["tpEvento"] = ev["tpEvento"].astype(str).str.strip()
            if "descEvento" in ev.columns:
                ev["descEvento"] = ev["descEvento"].astype(str).str.strip().str.lower()
            if "dhEvento" in ev.columns:
                ev["dhEvento"] = pd.to_datetime(ev["dhEvento"], errors="coerce")


            cancel_mask = pd.Series(False, index=ev.index)
            if "tpEvento" in ev.columns:
                cancel_mask |= ev["tpEvento"].eq("110111")
            if "descEvento" in ev.columns:
                cancel_mask |= ev["descEvento"].str.contains("cancel", na=False)
            ev_cancel = ev[cancel_mask].copy()

            if not ev_cancel.empty:
                ev_cancel = ev_cancel.sort_values(["__key", "dhEvento"], ascending=[True, True])
                last = ev_cancel.groupby("__key", as_index=False).tail(1)

                # montar lookup com arquivo do evento também
                keep_cols = ["__key", "_arquivo"]
                if "dhEvento" in last.columns:
                    keep_cols.append("dhEvento")
                if "nProt_retEvento" in last.columns:
                    keep_cols.append("nProt_retEvento")
                if "emit_CNPJ" in last.columns:   # <--- NOVO
                    keep_cols.append("emit_CNPJ")

                cancel_info = last[keep_cols].rename(columns={
                    "dhEvento": "cancelado_em",
                    "nProt_retEvento": "cancel_nProt",
                    "_arquivo": "_arquivo_evento"
                })

        # 2) Mover canceladas para "Erros"
        if not cancel_info.empty:
            # keys canceladas
            keys_cancel = set(cancel_info["__key"].dropna().astype(str))

            # a) NF reais canceladas (NF-e/NFC-e)
            mask_nf = df["_parser"].isin(["NF-e", "NFC-e"])
            nf_canceladas = df[mask_nf & df["__key"].isin(keys_cancel)].copy()

            if not nf_canceladas.empty:
                nf_canceladas = nf_canceladas.merge(cancel_info, on="__key", how="left")

                # normaliza tipos (garante que vão “inteiros” para o Excel)
                if "emissao" in nf_canceladas.columns:
                    nf_canceladas["emissao"] = pd.to_datetime(nf_canceladas["emissao"], errors="coerce")
                if "vNF" in nf_canceladas.columns:
                    nf_canceladas["vNF"] = nf_canceladas["vNF"].map(to_number_maybe_br)

                for _, r in nf_canceladas.iterrows():
                    erros.append({
                        "tipo": "NF cancelada",
                        "chave": r.get("chave") or normalize_key(r.get("chave")),
                        "nNF": r.get("nNF"),
                        "serie": r.get("serie"),
                        "emissao": r.get("emissao"),
                        "vNF": r.get("vNF"),
                        "emit_CNPJ": r.get("emit_CNPJ"),
                        "emit_xNome": r.get("emit_xNome"),
                        "dest_CNPJ": r.get("dest_CNPJ"),
                        "dest_xNome": r.get("dest_xNome"),
                        "cancelado_em": r.get("cancelado_em"),
                        "cancel_nProt": r.get("cancel_nProt"),
                        "_arquivo_nota": r.get("_arquivo"),
                        "_arquivo_evento": r.get("_arquivo_evento"),
                    })

                # remove da tabela principal
                df = df.drop(index=nf_canceladas.index, errors="ignore")
          
            # b) Cancelada SEM XML da NF (apenas evento)
            keys_nf = set(df.loc[df["_parser"].isin(["NF-e", "NFC-e"]), "__key"].dropna().astype(str))
            apenas_evento = sorted(keys_cancel - keys_nf)
            if apenas_evento:
                lk = cancel_info.set_index("__key")
                for k in apenas_evento:
                    emit_do_evento = lk.at[k, "emit_CNPJ"] if ("emit_CNPJ" in lk.columns and k in lk.index) else None
                    emit_da_chave  = cnpj_from_chave(k)
                    erros.append({
                        "tipo": "NF cancelada (sem XML da nota)",
                        "chave": k,
                        "nNF": None,
                        "serie": None,
                        "emissao": None,
                        "vNF": None,
                        "emit_CNPJ": emit_do_evento or emit_da_chave,  # << NOVO
                        "emit_xNome": None,  # evento não traz nome
                        "dest_CNPJ": None,
                        "dest_xNome": None,
                        "cancelado_em": lk.at[k, "cancelado_em"] if ("cancelado_em" in lk.columns and k in lk.index) else None,
                        "cancel_nProt": lk.at[k, "cancel_nProt"] if ("cancel_nProt" in lk.columns and k in lk.index) else None,
                        "_arquivo_nota": None,
                        "_arquivo_evento": lk.at[k, "_arquivo_evento"] if ("_arquivo_evento" in lk.columns and k in lk.index) else None,
                    })


            # c) opcional: remover eventos da tabela principal quando não quiser exibir
            # Remover eventos da tabela principal sempre (independente do checkbox)
            if df is not None and not df.empty and "_parser" in df.columns:
                df = df[df["_parser"] != "Evento NF-e"]


        # limpar coluna técnica
        if "__key" in df.columns:
            df.drop(columns=["__key"], inplace=True)

    # Salva no estado
    st.session_state.df = df
    st.session_state.df_view = None  # será montado abaixo
    st.session_state.erros = erros
    st.session_state.paths = paths

# ---------------------------
# Renderização usando o estado (sem reprocessar)
# ---------------------------
df = st.session_state.df
erros = st.session_state.erros
paths = st.session_state.paths


# Monta df_view (SEM canceladas e SEM eventos)
df_view = None
if df is not None and not df.empty:
    df_view = df.copy()

    # 1) Nunca mostrar eventos na visualização
    if "_parser" in df_view.columns:
        df_view = df_view[df_view["_parser"] != "Evento NF-e"]

    # 2) Nunca mostrar notas canceladas (essas ficam só na aba Erros)
    if "status_nota" in df_view.columns:
        df_view = df_view[df_view["status_nota"].ne("Cancelada")]

    # 3) (Opcional) limitar aos tipos de documento principais
    tipos_nota = {"NF-e","NFC-e","NFSe","NFSe RN (Prestado)","NFSe RN (Tomado)","CT-e","NF-e (sintético por evento)"}
    if "_parser" in df_view.columns:
        df_view = df_view[df_view["_parser"].isin(tipos_nota)]

    # 4) Colunas essenciais
    COLS_MIN = [
        "_parser","_arquivo","chave","nNF","serie","emissao",
        "emit_CNPJ","emit_xNome","dest_CNPJ","dest_xNome",
        "vNF","movimento","CFOPs_itens","CFOP_predominante",
        "vBC_ICMS","vICMS","vBC_ST","vICMS_ST",
    ]
    df_view = df_view[[c for c in COLS_MIN if c in df_view.columns]]




st.session_state.df_view = df_view

# Caso ainda não tenha rodado nada:
if df is None and (uploaded_files or dir_path.strip()) and not processar:
    st.info("Clique em **Processar** para iniciar a leitura.")
elif df is None:
    st.info("Envie arquivos ou informe o diretório e clique em **Processar**.")
else:
    # Saída em abas
    st.markdown("### 📊 Resultados")
    tabs = st.tabs(["Visualização", "Erros", "Exportar"])

    with tabs[0]:
        st.markdown(
            f"""
            <div class="az-card">
              <div>Arquivos únicos: <b>{len(paths)}</b> • Linhas totais: <b>{len(df)}</b> • Exibidas: <b>{0 if df_view is None else len(df_view)}</b> • Erros: <b>{0 if not erros else len(erros)}</b></div>
            </div>
            """,
            unsafe_allow_html=True
        )

        if df_view is not None and not df_view.empty:
            column_config = {}
            if "vNF" in df_view.columns:
                column_config["vNF"] = st.column_config.NumberColumn("vNF", help="Valor total da NF", format="R$ %.2f")
            if "emissao" in df_view.columns:
                column_config["emissao"] = st.column_config.DatetimeColumn("emissao", help="Data/hora de emissão", format="DD/MM/YYYY HH:mm:ss")

            st.dataframe(df_view.head(50), use_container_width=True, column_config=column_config)
        else:
            st.info("Nenhum registro a exibir.")

    with tabs[1]:
        st.markdown('<div class="az-card">', unsafe_allow_html=True)
        if erros:
            st.write("Ocorrências registradas (inclui **Notas Canceladas**):")
            df_err = pd.DataFrame(erros)

            # Ordena por tipo e chave para facilitar leitura
            order_cols = [c for c in ["tipo","chave","emissao","_arquivo_nota","_arquivo_evento"] if c in df_err.columns]
            if order_cols:
                df_err = df_err.sort_values(order_cols)

            st.dataframe(df_err, use_container_width=True)

            # Exportar ERROS em Excel (removendo timezone)
            out_err = io.BytesIO()
            with pd.ExcelWriter(out_err, engine="openpyxl") as writer:
                df_err_xl = strip_tz_for_excel(df_err.copy())

                # Normalizar colunas de data
                for dcol in ["emissao", "cancelado_em", "dhEvento"]:
                    if dcol in df_err_xl.columns:
                        df_err_xl[dcol] = pd.to_datetime(df_err_xl[dcol], errors="coerce")

                # Escreve no Excel
                df_err_xl.to_excel(writer, index=False, sheet_name="Erros")
                ws = writer.sheets["Erros"]

                # Formatação BR nas colunas de data
                for dcol in ["emissao", "cancelado_em", "dhEvento"]:
                    if dcol in df_err_xl.columns:
                        cidx = list(df_err_xl.columns).index(dcol) + 1
                        col_letter = get_column_letter(cidx)
                        for cell in ws[col_letter][1:]:  # pula cabeçalho
                            cell.number_format = "dd/mm/yyyy hh:mm:ss"

            st.download_button(
                "⬇️ Baixar erros (Excel)",
                data=out_err.getvalue(),
                file_name="erros_processamento.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_erros_excel",
            )

        else:
            st.info("Nenhum erro 🎉")
        st.markdown('</div>', unsafe_allow_html=True)

        with tabs[2]:
            st.markdown('<div class="az-card">', unsafe_allow_html=True)
            if df_view is None or df_view.empty:
                st.info("Nada para exportar.")
            else:
                # ===== Exportação Excel (único botão) =====
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    df_xl = strip_tz_for_excel(df_view).copy()
                    if "emissao" in df_xl.columns:
                        df_xl["emissao"] = pd.to_datetime(df_xl["emissao"], errors="coerce")
                    if "aliquota" in df_xl.columns:
                        df_xl["aliquota"] = pd.to_numeric(df_xl["aliquota"], errors="coerce")

                    colunas_novas = {
                        "vNF": "Valor",
                        "emissao": "Data de Emissão",
                        "emit_CNPJ": "CNPJ Emitente",
                        "emit_xNome": "Nome Emitente",
                        "dest_CNPJ": "CNPJ Destinatário",
                        "dest_xNome": "Nome Destinatário",
                        "nNF": "Número NF",
                        "serie": "Série",
                        "movimento": "Tipo (Entrada/Saída)",
                        "CFOPs_itens": "CFOP(s) da Nota",
                        "CFOP_predominante": "CFOP Predominante",
                        "vBC_ICMS": "BC ICMS",
                        "vICMS": "Valor ICMS",
                        "vBC_ST": "BC ICMS ST",
                        "vICMS_ST": "Valor ICMS ST",
                    }
                    df_xl = df_xl.rename(columns=colunas_novas)
                    df_xl.to_excel(writer, index=False, sheet_name="Notas")
                    ws = writer.sheets["Notas"]

                    from openpyxl.utils import get_column_letter
                    def fmt_col(nome_coluna: str, number_format: str):
                        if nome_coluna in df_xl.columns:
                            cidx = list(df_xl.columns).index(nome_coluna) + 1
                            col_letter = get_column_letter(cidx)
                            for cell in ws[col_letter][1:]:
                                cell.number_format = number_format

                    fmt_col("Valor", 'R$ #,##0.00')
                    fmt_col("Data de Emissão", 'dd/mm/yyyy hh:mm:ss')
                    fmt_col("BC ICMS", 'R$ #,##0.00')
                    fmt_col("Valor ICMS", 'R$ #,##0.00')
                    fmt_col("BC ICMS ST", 'R$ #,##0.00')
                    fmt_col("Valor ICMS ST", 'R$ #,##0.00')

                st.download_button(
                    "⬇️ Baixar Excel",
                    data=out.getvalue(),
                    file_name="notas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_notas_excel",
                )
    

        st.markdown('</div>', unsafe_allow_html=True)

    st.caption(
        "Entrada/Saída por tpNF; eventos de cancelamento (110111) são enviados para a aba **Erros** e removidos da tabela principal."
    )
